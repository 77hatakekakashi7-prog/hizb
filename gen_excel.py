import json, sys, base64, io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference

def sd(style='thin', color='BBBBBB'):
    return Side(style=style, color=color)

def bdr(t='thin',b='thin',l='thin',r='thin',tc='BBBBBB',bc='BBBBBB',lc='BBBBBB',rc='BBBBBB'):
    return Border(top=sd(t,tc),bottom=sd(b,bc),left=sd(l,lc),right=sd(r,rc))

THIN  = bdr()
MED_G = bdr('medium','medium','medium','medium','9B7210','9B7210','9B7210','9B7210')
INC_B = bdr('medium','medium','medium','medium','00A550','00A550','00A550','00A550')
EXP_B = bdr('medium','medium','medium','medium','C1121F','C1121F','C1121F','C1121F')
SUM_B = bdr('medium','medium','medium','medium','1E1E36','1E1E36','1E1E36','1E1E36')

def fl(h): return PatternFill('solid', fgColor=h)
def fn(sz=10,bold=False,color='000000',italic=False): return Font(name='Arial',size=sz,bold=bold,color=color,italic=italic)
def al(h='left',v='center',wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap,readingOrder=2)

NUM_FMT = '#,##0.00 "EGP"'

def sc(c,val=None,sz=10,bold=False,color='000000',italic=False,bg=None,h='left',border=None,num_fmt=None):
    if val is not None: c.value=val
    c.font=fn(sz,bold,color,italic)
    if bg: c.fill=fl(bg)
    c.alignment=al(h)
    if border: c.border=border
    if num_fmt: c.number_format=num_fmt

def build_workbook(txs, period_label='تقرير مالي'):
    wb = Workbook()
    wb.remove(wb.active)

    inc_total = sum(t['amt'] for t in txs if t['type']=='income')
    exp_total = sum(t['amt'] for t in txs if t['type']=='expense')
    bal = inc_total - exp_total

    cats = {}
    for t in txs:
        cats.setdefault(t['cat'],{'inc':0,'exp':0,'n':0})
        if t['type']=='income': cats[t['cat']]['inc']+=t['amt']
        else: cats[t['cat']]['exp']+=t['amt']
        cats[t['cat']]['n']+=1

    # ── SHEET 1: العمليات ────────────────────────────────────────────
    ws1 = wb.create_sheet('العمليات')
    ws1.sheet_view.rightToLeft = True
    ws1.sheet_properties.tabColor = '7A0012'
    ws1.freeze_panes = 'A5'

    ws1.row_dimensions[1].height = 38
    ws1.merge_cells('A1:G1')
    sc(ws1['A1'], f'★  الحزب الاشتراكى — {period_label}  ★', sz=15, bold=True, color='FFFFFF', bg='7A0012', h='center', border=MED_G)

    ws1.row_dimensions[2].height = 16
    ws1.merge_cells('A2:G2')
    sc(ws1['A2'], f'تاريخ التقرير: {datetime.now().strftime("%Y-%m-%d  %H:%M")}', sz=9, italic=True, color='7A6510', bg='F5EFD4', h='center')

    ws1.row_dimensions[3].height = 5
    for col in range(1,8): ws1.cell(3,col).fill=fl('F0EBD5')

    ws1.row_dimensions[4].height = 22
    for ci,h in enumerate(['التاريخ','النوع','الوصف','الفئة','المبلغ (EGP)','ملاحظة','الحالة'],1):
        sc(ws1.cell(4,ci), h, bold=True, color='FFFFFF', bg='1E1E36', h='center', border=THIN)

    sorted_txs = sorted(txs, key=lambda t: t['dt'])
    for ri,t in enumerate(sorted_txs):
        row=5+ri; ws1.row_dimensions[row].height=18
        is_inc = t['type']=='income'
        bg = ('C6EFCE' if ri%2==0 else 'B0E0BC') if is_inc else ('FFC7CE' if ri%2==0 else 'FFAABB')
        fc = '1A6B35' if is_inc else '7A0012'
        brd = INC_B if is_inc else EXP_B
        amt = t['amt'] if is_inc else -t['amt']
        vals=[t['dt'],'▲ دخل' if is_inc else '▼ مصروف',t['dsc'],t['cat'],amt,t.get('note',''),'✔ مكتمل']
        aligns=['center','center','left','center','right','left','center']
        for ci,(v,h_al) in enumerate(zip(vals,aligns),1):
            c=ws1.cell(row,ci,v); c.font=fn(10,bold=(ci==5),color=fc)
            c.fill=fl(bg); c.alignment=al(h_al); c.border=brd
            if ci==5: c.number_format=NUM_FMT

    data_end = 4+len(sorted_txs)
    if sorted_txs:
        tbl=Table(displayName='TxTable',ref=f'A4:G{data_end}')
        tbl.tableStyleInfo=TableStyleInfo(name='TableStyleMedium9',showFirstColumn=False,showLastColumn=False,showRowStripes=False,showColumnStripes=False)
        ws1.add_table(tbl)

    for i,w in enumerate([13,12,32,20,18,28,11],1):
        ws1.column_dimensions[get_column_letter(i)].width=w

    ws1.page_setup.orientation='landscape'; ws1.page_setup.fitToPage=True; ws1.page_setup.fitToWidth=1
    ws1.print_title_rows='1:4'

    # ── SHEET 2: الملخص ─────────────────────────────────────────────
    ws2 = wb.create_sheet('الملخص')
    ws2.sheet_view.rightToLeft = True
    ws2.sheet_properties.tabColor = 'D4A017'

    ws2.row_dimensions[1].height=36; ws2.merge_cells('A1:E1')
    sc(ws2['A1'],f'★  الملخص المالي — {period_label}  ★',sz=14,bold=True,color='FFFFFF',bg='7A0012',h='center',border=MED_G)
    ws2.row_dimensions[2].height=16; ws2.merge_cells('A2:E2')
    sc(ws2['A2'],f'تاريخ التقرير: {datetime.now().strftime("%Y-%m-%d")}',sz=9,italic=True,color='7A6510',bg='F5EFD4',h='center')

    ws2.row_dimensions[4].height=22; ws2.row_dimensions[5].height=36
    kpis=[
        ('إجمالي الدخل',inc_total,'C6EFCE','1A6B35','00A550'),
        ('إجمالي المصروف',-exp_total,'FFC7CE','7A0012','C1121F'),
        ('صافي الرصيد',bal,('FFF2CC' if bal>=0 else 'FFC7CE'),('7A6510' if bal>=0 else '7A0012'),('D4A017' if bal>=0 else 'C1121F')),
        ('عدد العمليات',len(txs),'EDE8D8','1E1E36','555555'),
    ]
    for ci,(label,val,bg,fg,bc) in enumerate(kpis,1):
        hb=bdr('medium','thin','medium','medium',bc,'BBBBBB',bc,bc)
        vb=bdr('thin','medium','medium','medium','BBBBBB',bc,bc,bc)
        sc(ws2.cell(4,ci),label,bold=True,color='FFFFFF',bg='2E2E50',h='center',border=hb)
        c=ws2.cell(5,ci,val); c.font=fn(14,bold=True,color=fg); c.fill=fl(bg)
        c.alignment=al('center'); c.border=vb
        if ci<=3: c.number_format=NUM_FMT

    ws2.row_dimensions[7].height=22
    for ci,h in enumerate(['الفئة','العمليات','مصروف (EGP)','دخل (EGP)','صافي (EGP)'],1):
        sc(ws2.cell(7,ci),h,bold=True,color='FFFFFF',bg='4A3A00',h='center',border=THIN)

    cat_start=8
    for ri2,(cat,v) in enumerate(sorted(cats.items())):
        row2=cat_start+ri2; ws2.row_dimensions[row2].height=18
        net=v['inc']-v['exp']; alt='F5F0DC' if ri2%2==0 else 'FFFFFF'
        rows_data=[
            (1,cat,'left',alt,'1E1E36',False),
            (2,v['n'],'center',alt,'1E1E36',False),
            (3,-v['exp'] if v['exp'] else 0,'right','FFF0F0' if v['exp'] else alt,'7A0012' if v['exp'] else '888888',True),
            (4,v['inc'] if v['inc'] else 0,'right','F0FFF4' if v['inc'] else alt,'1A6B35' if v['inc'] else '888888',True),
            (5,net,'right',('FFF2CC' if net>0 else ('FFF0F0' if net<0 else alt)),('7A6510' if net>0 else ('7A0012' if net<0 else '888888')),True),
        ]
        for ci,val,h_al,bg,fg,is_num in rows_data:
            c=ws2.cell(row2,ci,val); c.font=fn(10,color=fg); c.fill=fl(bg)
            c.alignment=al(h_al); c.border=THIN
            if is_num: c.number_format=NUM_FMT

    cat_end=cat_start+len(cats)-1
    if cats:
        tr=cat_end+1; ws2.row_dimensions[tr].height=20
        for ci,val in [(1,'الإجمالي'),(2,len(txs)),(3,-exp_total),(4,inc_total),(5,bal)]:
            c=ws2.cell(tr,ci,val); c.font=fn(10,bold=True,color='FFFFFF')
            c.fill=fl('1E1E36'); c.alignment=al('right' if ci>1 else 'center'); c.border=SUM_B
            if ci>2: c.number_format=NUM_FMT
        tbl2=Table(displayName='CatTable',ref=f'A7:E{cat_end}')
        tbl2.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showFirstColumn=True,showLastColumn=False,showRowStripes=False,showColumnStripes=False)
        ws2.add_table(tbl2)

    for i,w in enumerate([22,14,18,18,18],1):
        ws2.column_dimensions[get_column_letter(i)].width=w

    ws2.page_setup.orientation='landscape'; ws2.page_setup.fitToPage=True; ws2.page_setup.fitToWidth=1

    # ── SHEET 3: رسم بياني ──────────────────────────────────────────
    ws3 = wb.create_sheet('رسم بياني')
    ws3.sheet_view.rightToLeft = True
    ws3.sheet_properties.tabColor = '22C55E'

    for ci,h in enumerate(['الفئة','دخل','مصروف'],1):
        sc(ws3.cell(1,ci),h,bold=True,color='FFFFFF',bg='1E1E36',h='center',border=THIN)

    for ri3,(cat,v) in enumerate(sorted(cats.items()),2):
        ws3.row_dimensions[ri3].height=17; alt='F5F0DC' if ri3%2==0 else 'FFFFFF'
        c1=ws3.cell(ri3,1,cat); c1.fill=fl(alt); c1.border=THIN; c1.alignment=al('left')
        c2=ws3.cell(ri3,2,v['inc']); c2.font=fn(10,color='1A6B35'); c2.fill=fl(alt); c2.border=THIN; c2.alignment=al('right'); c2.number_format=NUM_FMT
        c3=ws3.cell(ri3,3,v['exp']); c3.font=fn(10,color='7A0012'); c3.fill=fl(alt); c3.border=THIN; c3.alignment=al('right'); c3.number_format=NUM_FMT

    last_data=1+len(cats)
    if cats:
        chart=BarChart(); chart.type='col'
        chart.title='دخل ومصروف حسب الفئة'; chart.style=10
        chart.y_axis.title='المبلغ (EGP)'; chart.x_axis.title='الفئة'
        chart.width=22; chart.height=14
        dr=Reference(ws3,min_col=2,max_col=3,min_row=1,max_row=last_data)
        cr=Reference(ws3,min_col=1,min_row=2,max_row=last_data)
        chart.add_data(dr,titles_from_data=True); chart.set_categories(cr)
        chart.series[0].graphicalProperties.solidFill='22C55E'
        chart.series[1].graphicalProperties.solidFill='C1121F'
        ws3.add_chart(chart,'E2')

    for i,w in enumerate([22,18,18],1):
        ws3.column_dimensions[get_column_letter(i)].width=w

    ws3.page_setup.orientation='landscape'; ws3.page_setup.fitToPage=True; ws3.page_setup.fitToWidth=1

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

if __name__=='__main__':
    txs=json.loads(sys.argv[1] if len(sys.argv)>1 else '[]')
    label=sys.argv[2] if len(sys.argv)>2 else 'تقرير مالي'
    sys.stdout.buffer.write(base64.b64encode(build_workbook(txs,label)))
